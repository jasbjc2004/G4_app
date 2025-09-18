import os

import openpyxl
import pandas as pd
from PySide6.QtCore import QThread, Signal, QMutex, QWaitCondition, Qt
from PySide6.QtGui import QTextCursor, QColor
from openpyxl.styles import Alignment, Font

from constants import UNIMAN_PARAMS, BIMAN_PARAMS, LETTER_SIZE, SUBTITLE_LETTER_SIZE, SUB_SUB_TITLE_LETTER_SIZE, \
    FONT_LETTER_SIZE
from data_processing import calculate_extra_parameters
from widget_settings import manage_settings


def make_time(sec: float):
    """
    Convert a certain amount of seconds to the 'min:sec'-format
    :param sec:
    :return:
    """
    minit = int(sec // 60)
    sec = sec % 60
    return f"{minit:02d}:{sec:06.3f}"


class DownloadThread(QThread):
    progress = Signal(int)
    pdf_ready_image = Signal(int, list, list, list, list, list, tuple)
    finished_file = Signal()
    error_occurred = Signal(str)

    def __init__(self, parent, part_folder, index=-1, id=None, pdf=None, check=None):
        """
        Download all the information of the participant to both PDF and Excel, or only one Excel (if index is -1)
        Makes it possible to have a progression bar and exports all calculation to a separate thread
        Needed to let the mainthread handle all the relavant windows and updating of plot (otherwise not possible)
        :param parent: parent window to collect the data
        :type parent: MainWindow
        :param part_folder: participant folder to save the data
        :param index: which trial to download (if index is not -1), otherwise download all info
        :param pdf: the PDF to write the information to
        :param check: all the necessary plots that need to be added to the PDF
        """
        super().__init__()

        self.gopro_time = []
        self.pdf = pdf
        self.participant_folder = part_folder
        self.checkboxes = check

        self.total_num_trials = parent.num_trials
        self.num_trials = parent.num_trials
        self.main = parent
        self.counter_progress = 0

        active_tabs = self.count_active_tabs()
        if active_tabs == 0: active_tabs = 100
        if self.checkboxes is not None and len(self.checkboxes) == 0:
            self.step_progress = 100 / active_tabs
        elif self.checkboxes is not None:
            self.step_progress = 100 / (active_tabs * len(self.checkboxes))
        else:
            self.step_progress = 100

        self.part_id = id

        self.mutex = QMutex()
        self.condition = QWaitCondition()

        self.index = index

    def count_active_tabs(self):
        """
        Count all the active tabs, with data
        :rtype: int
        """
        from widget_trials import TrailTab

        counter_active = 0
        for i in range(self.num_trials):
            tab = self.main.tab_widget.widget(i)

            if isinstance(tab, TrailTab) and len(tab.xs) > 0:
                counter_active += 1
        return counter_active

    def run(self):
        try:
            if self.pdf:
                self.num_trials = self.count_active_tabs()
                self.pdf.cell(0, 8, f"Total trials: {self.total_num_trials}", ln=True)
                self.pdf.cell(0, 8, f"Total used trials: {self.num_trials}", ln=True)

            range_index = list(range(self.total_num_trials)) if self.index == -1 else [self.index]
            for i in range_index:
                self.export_tab(i)

            if self.pdf:
                self.pdf.add_page()
                self.pdf.set_font("Arial", style="B", size=SUBTITLE_LETTER_SIZE)
                self.pdf.cell(0, 10, f"Average over all trials with score 3", ln=True)
                self.pdf.set_font("Arial", size=LETTER_SIZE)

                self.average_events_info()

                self.final_excel()

            self.finished_file.emit()

        except Exception as e:
            self.error_occurred.emit(str(e))

    def export_tab(self, index):
        """
        Export the information of trial at the corresponding index
        """
        from widget_trials import TrailTab

        tab = self.main.tab_widget.widget(index)

        if isinstance(tab, TrailTab):
            NUMBER_EVENTS = manage_settings.get("Events", "NUMBER_EVENTS")

            events = [ei if ei is not None else 0 for ei in tab.event_log[0:NUMBER_EVENTS]]
            self.gopro_time = [round(tab.xs[events[i]] + tab.trial_time_start, 3)
                               if (tab.trial_time_start != 0 and events[i] != 0) else
                               round(tab.trial_time_start, 3) for i in range(NUMBER_EVENTS)]
            print(self.gopro_time)

            print('done making time')

            data = {
                "Time (s)": tab.xs if tab.xs else [],
                "Left Sensor x (cm)": [pos[0] for pos in tab.log_left] if tab.xs else [],
                "Left Sensor y (cm)": [pos[1] for pos in tab.log_left] if tab.xs else [],
                "Left Sensor z (cm)": [pos[2] for pos in tab.log_left] if tab.xs else [],
                "Left Sensor v (m/s)": [pos[3] for pos in tab.log_left] if tab.xs else [],
                "Right Sensor x (cm)": [pos[0] for pos in tab.log_right] if tab.xs else [],
                "Right Sensor y (cm)": [pos[1] for pos in tab.log_right] if tab.xs else [],
                "Right Sensor z (cm)": [pos[2] for pos in tab.log_right] if tab.xs else [],
                "Right Sensor v (m/s)": [pos[3] for pos in tab.log_right] if tab.xs else [],
                "Score:": [tab.get_score()] if tab.xs else [],
                " ": [],
                "Automatic events (/):": [tab.event_log[i] if tab.event_old_log[i] == 0 else tab.event_old_log[i] for i
                                          in range(NUMBER_EVENTS)] if tab.xs else [],
                "Manual events (/):": (
                    [0] * NUMBER_EVENTS if all(e == 0 for e in tab.event_old_log) else tab.event_log) if tab.xs else [],
                "Position events:": tab.event_position if tab.xs else [],
                "": [],
                "Events (s)": [tab.xs[tab.event_log[i]] if tab.event_old_log[i] == 0 else tab.event_old_log[i] for i in
                               range(NUMBER_EVENTS)] if tab.xs else [],
                "GoPro events (s):": self.gopro_time,
                "GoPro start (s)": [tab.trial_time_start],
            }
            max_length = max(len(v) for v in data.values())
            for key in data:
                data[key].extend([None] * (max_length - len(data[key])))
            df = pd.DataFrame(data)
            trial_file = os.path.join(self.participant_folder, f"trial_{index + 1}.xlsx")
            print('done making excel')
            if os.path.exists(trial_file):
                os.remove(trial_file)
            with pd.ExcelWriter(trial_file) as writer:
                df.to_excel(writer, index=False)
            if self.pdf:
                current_y = self.pdf.get_y()
                page_height = self.pdf.h - 20  # margin
                available_space = page_height - current_y

                if available_space < 40:
                    self.pdf.add_page()
                    self.y_image = self.pdf.get_y()

                self.pdf.set_font("Arial", style="B", size=SUBTITLE_LETTER_SIZE)
                if tab.case_status in [0, 5]:
                    box_hand = 'Left'
                elif tab.case_status in [1, 4]:
                    box_hand = 'Right'
                else:
                    box_hand = 'Both'
                self.pdf.cell(0, 10,
                              f"Trial {index + 1}:{f' score {tab.get_score()} & Box Hand: {box_hand}' if tab.xs else ''}",
                              ln=True)
                self.pdf.set_font("Arial", size=LETTER_SIZE)

                doc = tab.notes_input.document()
                block = doc.begin()

                black_fragments = []
                red_fragments = []

                while block.isValid():
                    cursor = QTextCursor(block)
                    if cursor.currentTable() is None:
                        fmt = cursor.charFormat()
                        color = fmt.foreground().color()
                        text = block.text()

                        if color == QColor(Qt.red):
                            red_fragments.append(text)
                        else:
                            black_fragments.append(text)

                    block = block.next()

                filtered_red_fragments = [fram for fram in red_fragments if fram.strip() != '']
                filtered_black_fragments = [fram for fram in black_fragments if fram.strip() != '']

                if filtered_red_fragments:
                    red_text = "\n".join(filtered_red_fragments)
                else:
                    red_text = "No Automatic Notes"
                if filtered_black_fragments:
                    black_text = "\n".join(filtered_black_fragments)
                else:
                    black_text = "No Additional Notes"

                self.pdf.multi_cell(0, 6, black_text)

                self.pdf.ln(5)

                self.pdf.set_text_color(255, 0, 0)
                self.pdf.multi_cell(0, 6, red_text)
                self.pdf.set_text_color(0, 0, 0)

                self.pdf.ln(5)

                if tab.xs:
                    self.pdf.set_font("Arial", style="B", size=SUB_SUB_TITLE_LETTER_SIZE)
                    self.pdf.cell(0, 8, 'Events', ln=True)

                    col_widths = [8, 38, 15, 30, 30, 35]
                    self.pdf.set_font('Arial', '', FONT_LETTER_SIZE)
                    if tab.trial_time_start != 0:
                        self.pdf.cell(0, 6, f'Starting the trial at: {make_time(round(tab.trial_time_start, 3))}', ln=True)

                    events_table = [
                        ['', '', 'Frame', 'Absolute time (s)', 'Relative time (s)', 'GoPro time (min:sec)'],
                        ['e1', 'Start BH', events[0], round(tab.xs[events[0]], 2), 0, make_time(self.gopro_time[0])],
                        ['e2', 'Start box opening', events[1], round(tab.xs[events[1]], 2),
                         round(tab.xs[events[1]] - tab.xs[events[0]], 2), make_time(self.gopro_time[1])],
                        ['e3', 'End box opening', events[2], round(tab.xs[events[2]], 2),
                         round(tab.xs[events[2]] - tab.xs[events[0]], 2), make_time(self.gopro_time[2])],
                        ['e4', 'Anticipation TH', events[3], round(tab.xs[events[3]], 2),
                         round(tab.xs[events[3]] - tab.xs[events[0]], 2), make_time(self.gopro_time[3])],
                        ['e5', 'Start movement to trigger', events[4], round(tab.xs[events[4]], 2)
                        if tab.xs[events[3]] != tab.xs[events[4]] else '',
                         round(tab.xs[events[4]] - tab.xs[events[0]], 2)
                         if tab.xs[events[3]] != tab.xs[events[4]] else '', make_time(self.gopro_time[4])],
                        ['e6', 'End of trial', events[5], round(tab.xs[events[5]], 2),
                         round(tab.xs[events[5]] - tab.xs[events[0]], 2), make_time(self.gopro_time[5])]
                    ]
                    events_table = [[str(cell) if cell != '' else '' for cell in row] for row in events_table]

                    line_height = 8
                    self.pdf.set_fill_color(235, 235, 235)  # lichtgrijs
                    self.pdf.set_text_color(0, 0, 0)

                    for row_ind, row in enumerate(events_table):
                        self.pdf.set_x(20)
                        first_row = (row_ind == 0)

                        for col_ind, datum in enumerate(row):
                            fill = first_row or col_ind == 0
                            align = 'L' if col_ind in [0, 1] and row_ind != 0 else 'C'
                            style = 'B' if first_row or col_ind == 0 else ''
                            self.pdf.set_font("Arial", style, size=8)
                            self.pdf.cell(col_widths[col_ind], line_height, datum, border=1, align=align, fill=fill)

                        self.pdf.ln(line_height)

                    self.pdf.ln(5)

                    count_imag = 0
                    for pos_index in self.checkboxes:
                        left_data = [data["Left Sensor x (cm)"] if pos_index == 0 else
                                     data["Left Sensor y (cm)"] if pos_index == 1 else
                                     data["Left Sensor z (cm)"] if pos_index == 2 else
                                     data["Left Sensor v (m/s)"]][0]
                        right_data = [data["Right Sensor x (cm)"] if pos_index == 0 else
                                      data["Right Sensor y (cm)"] if pos_index == 1 else
                                      data["Right Sensor z (cm)"] if pos_index == 2 else
                                      data["Right Sensor v (m/s)"]][0]
                        events = [ei if ei is not None else 0 for ei in tab.event_log]

                        self.mutex.lock()
                        self.pdf_ready_image.emit(pos_index, tab.xs, left_data, right_data, events,
                                                  tab.event_position, (count_imag, len(self.checkboxes)))
                        count_imag += 1
                        self.condition.wait(self.mutex)
                        self.mutex.unlock()

                        self.counter_progress += self.step_progress
                        self.progress.emit(round(self.counter_progress))

                if self.counter_progress < (index + 1) * self.step_progress:
                    self.counter_progress += self.step_progress
                    self.progress.emit(round(self.counter_progress))

                if tab.get_score() == 3:
                    events = [ei if ei is not None else 0 for ei in tab.event_log[0:NUMBER_EVENTS]]
                    tab.extra_parameters_bim, tab.extra_parameters_uni = calculate_extra_parameters(events,
                                                                                                    tab.log_left,
                                                                                                    tab.log_right)

                    self.pdf.set_font("Arial", style="B", size=9)
                    self.pdf.cell(0, 10, 'Parameters', ln=True)

                    col_widths = [45, 60, 40]
                    self.pdf.set_font('Arial', '', 10)

                    data = [
                        ['', 'Parameter', 'Value'],
                        ['Bimanual', BIMAN_PARAMS[0], str(round(tab.extra_parameters_bim[0], 2))],
                        ['', BIMAN_PARAMS[1], str(round(tab.extra_parameters_bim[1], 2))],
                        ['', BIMAN_PARAMS[2], str(round(tab.extra_parameters_bim[2], 2))],
                        ['', BIMAN_PARAMS[3], str(round(tab.extra_parameters_bim[3], 2))],

                        ['Unimanual', UNIMAN_PARAMS[0], str(round(tab.extra_parameters_uni[0], 2))],
                        ['', UNIMAN_PARAMS[1], str(round(tab.extra_parameters_uni[1], 2))],
                        ['', UNIMAN_PARAMS[2], str(round(tab.extra_parameters_uni[2], 2))],
                        ['', UNIMAN_PARAMS[3], str(round(tab.extra_parameters_uni[3], 2))],
                        ['', UNIMAN_PARAMS[4], str(round(tab.extra_parameters_uni[4], 2))],
                        ['', UNIMAN_PARAMS[5], str(round(tab.extra_parameters_uni[5], 2))],
                        ['', UNIMAN_PARAMS[6], str(round(tab.extra_parameters_uni[6], 2))],
                        ['', UNIMAN_PARAMS[7], str(round(tab.extra_parameters_uni[7], 2))],
                        ['', UNIMAN_PARAMS[8], str(round(tab.extra_parameters_uni[8], 2))],
                        ['', UNIMAN_PARAMS[9], str(round(tab.extra_parameters_uni[9], 2))],
                    ]

                    line_height = 8
                    self.pdf.set_fill_color(235, 235, 235)  # lichtgrijs
                    self.pdf.set_text_color(0, 0, 0)

                    for row_ind, row in enumerate(data):
                        self.pdf.set_x(20)
                        first_row = (row_ind == 0)

                        for col_ind, datum in enumerate(row):
                            fill = first_row or col_ind == 0
                            align = 'L' if col_ind in [0, 1] and row_ind != 0 else 'C'
                            style = 'B' if first_row or col_ind == 0 else ''
                            self.pdf.set_font("Arial", style, size=8)
                            self.pdf.cell(col_widths[col_ind], line_height, datum, border=1, align=align, fill=fill)

                        self.pdf.ln(line_height)

    def average_events_info(self):
        from widget_trials import TrailTab

        average_bim_left = [0] * 4
        average_uni_left = [0] * 10
        left_counter = 0
        average_bim_right = [0] * 4
        average_uni_right = [0] * 10
        right_counter = 0

        for i in range(self.main.tab_widget.count()):
            tab = self.main.tab_widget.widget(i)

            if isinstance(tab, TrailTab) and len(tab.extra_parameters_bim) > 0:
                if tab.case_status == 0 and average_bim_left[0] == 0:
                    average_bim_left = tab.extra_parameters_bim
                    average_uni_left = tab.extra_parameters_uni
                    left_counter = 1
                elif tab.case_status == 1 and average_bim_right[0] == 0:
                    average_bim_right = tab.extra_parameters_bim
                    average_uni_right = tab.extra_parameters_uni
                    right_counter = 1
                elif tab.case_status == 0:
                    average_bim_left = tuple(
                        [a + b for a, b in zip(average_bim_left, tab.extra_parameters_bim)])
                    average_uni_left = tuple(
                        [a + b for a, b in zip(average_uni_left, tab.extra_parameters_uni)])
                    left_counter += 1
                elif tab.case_status == 1:
                    average_bim_right = tuple(
                        [a + b for a, b in zip(average_bim_right, tab.extra_parameters_bim)])
                    average_uni_right = tuple(
                        [a + b for a, b in zip(average_uni_right, tab.extra_parameters_uni)])
                    right_counter += 1

        average_bim_left = tuple([temp / left_counter if left_counter != 0 else 0 for temp in average_bim_left])
        average_uni_left = tuple([temp / left_counter if left_counter != 0 else 0 for temp in average_uni_left])
        average_bim_right = tuple([temp / right_counter if right_counter != 0 else 0 for temp in average_bim_right])
        average_uni_right = tuple([temp / right_counter if right_counter != 0 else 0 for temp in average_uni_right])

        col_widths = [45, 60, 40, 40]
        self.pdf.set_font('Arial', '', 11)
        self.pdf.cell(0, 6, f"Total used trials: {self.count_active_tabs()}", ln=True)

        data = [
            ['', 'Parameter', 'Average left (BH)', 'Average right (BH)'],
            ['', 'Total trials', str(left_counter), str(right_counter)],
            ['Bimanual', BIMAN_PARAMS[0], str(round(average_bim_left[0], 2)), str(round(average_bim_right[0], 2))],
            ['', BIMAN_PARAMS[1], str(round(average_bim_left[1], 2)), str(round(average_bim_right[1], 2))],
            ['', BIMAN_PARAMS[2], str(round(average_bim_left[2], 2)), str(round(average_bim_right[2], 2))],
            ['', BIMAN_PARAMS[3], str(round(average_bim_left[3], 2)), str(round(average_bim_right[3], 2))],

            ['Unimanual', UNIMAN_PARAMS[0], str(round(average_uni_left[0], 2)), str(round(average_uni_right[0], 2))],
            ['', UNIMAN_PARAMS[1], str(round(average_uni_left[1], 2)), str(round(average_uni_right[1], 2))],
            ['', UNIMAN_PARAMS[2], str(round(average_uni_left[2], 2)), str(round(average_uni_right[2], 2))],
            ['', UNIMAN_PARAMS[3], str(round(average_uni_left[3], 2)), str(round(average_uni_right[3], 2))],
            ['', UNIMAN_PARAMS[4], str(round(average_uni_left[4], 2)), str(round(average_uni_right[4], 2))],
            ['', UNIMAN_PARAMS[5], str(round(average_uni_left[5], 2)), str(round(average_uni_right[5], 2))],
            ['', UNIMAN_PARAMS[6], str(round(average_uni_left[6], 2)), str(round(average_uni_right[6], 2))],
            ['', UNIMAN_PARAMS[7], str(round(average_uni_left[7], 2)), str(round(average_uni_right[7], 2))],
            ['', UNIMAN_PARAMS[8], str(round(average_uni_left[8], 2)), str(round(average_uni_right[8], 2))],
            ['', UNIMAN_PARAMS[9], str(round(average_uni_left[9], 2)), str(round(average_uni_right[9], 2))],
        ]

        line_height = 8
        self.pdf.set_fill_color(235, 235, 235)
        self.pdf.set_text_color(0, 0, 0)

        for row_ind, row in enumerate(data):
            self.pdf.set_x(15)
            first_row = (row_ind == 0)

            for col_ind, datum in enumerate(row):
                fill = first_row or col_ind == 0
                align = 'L' if col_ind in [0, 1] and row_ind != 0 else 'C'
                style = 'B' if first_row or col_ind == 0 else ''
                self.pdf.set_font("Arial", style, size=LETTER_SIZE)
                self.pdf.cell(col_widths[col_ind], line_height, datum, border=1, align=align, fill=fill)

            self.pdf.ln(line_height)

    def final_excel(self):
        """
        Make a summary out of all the data, contains the events and bimanual and unimanual parameters
        :return:
        """
        LABEL_EVENT = manage_settings.get("Events", "LABEL_EVENT")

        range_index = list(range(self.total_num_trials)) if self.index == -1 else [self.index]

        sum_data = {
            "Events": {}
        }
        aver_data = {
            "": {"Number of Trials": [0, 0]}
        }
        for key in LABEL_EVENT:
            sum_data["Events"][key] = []

        sum_data[" "] = {}
        sum_data[" "][" "] = []
        sum_data[" "]["Score"] = []
        sum_data[" "][""] = []

        sum_data["Bimanual"] = {}
        aver_data["Bimanual"] = {}
        for param in BIMAN_PARAMS:
            sum_data["Bimanual"][param] = []
            aver_data["Bimanual"][param] = [0, 0]

        sum_data["Unimanual"] = {}
        aver_data["Unimanual"] = {}
        for param in UNIMAN_PARAMS:
            sum_data["Unimanual"][param] = []
            aver_data["Unimanual"][param] = [0, 0]

        valid_ranges = [index for index in range_index if self.main.tab_widget.widget(index).xs]
        print(range_index, valid_ranges)
        for index in valid_ranges:
            print(index)
            tab = self.main.tab_widget.widget(index)

            sum_data[" "]["Score"].append(tab.get_score())

            for i, ei in enumerate(tab.event_log):
                sum_data["Events"][LABEL_EVENT[i]].append(tab.xs[ei])

            sum_data[" "][" "].append(" ")
            sum_data[" "][""].append(" ")

            for i, param in enumerate(tab.extra_parameters_bim):
                sum_data["Bimanual"][BIMAN_PARAMS[i]].append(param)

                if tab.case_status == 0:
                    aver_data["Bimanual"][BIMAN_PARAMS[i]][0] += param
                elif tab.case_status == 1:
                    aver_data["Bimanual"][BIMAN_PARAMS[i]][1] += param

            if tab.case_status == 0:
                aver_data[""]["Number of Trials"][0] += 1
            elif tab.case_status == 1:
                aver_data[""]["Number of Trials"][1] += 1

            for i, param in enumerate(tab.extra_parameters_uni):
                sum_data["Unimanual"][UNIMAN_PARAMS[i]].append(param)

                if tab.case_status == 0:
                    aver_data["Unimanual"][UNIMAN_PARAMS[i]][0] += param
                elif tab.case_status == 1:
                    aver_data["Unimanual"][UNIMAN_PARAMS[i]][1] += param

        for param in aver_data["Bimanual"]:
            aver_data["Bimanual"][param] = [param_lr / aver_data[""]["Number of Trials"][index]
                                            if aver_data[""]["Number of Trials"][index] != 0 else 0
                                            for index, param_lr in enumerate(aver_data["Bimanual"][param])]

        for param in aver_data["Unimanual"]:
            aver_data["Unimanual"][param] = [param_lr / aver_data[""]["Number of Trials"][index]
                                             if aver_data[""]["Number of Trials"][index] != 0 else 0
                                             for index, param_lr in enumerate(aver_data["Unimanual"][param])]

        columns_sum = pd.MultiIndex.from_tuples([(heading, sub) for heading, subdict in sum_data.items()
                                                 for sub in subdict])
        first_heading = next(iter(sum_data))
        first_sub = next(iter(sum_data[first_heading]))
        n_rows = len(sum_data[first_heading][first_sub])
        data_sum = [
            [sum_data[heading][sub][i] for heading, subdict in sum_data.items() for sub in subdict]
            for i in range(n_rows)
        ]
        df_sum = pd.DataFrame(data_sum, columns=columns_sum)

        columns_aver = pd.MultiIndex.from_tuples(
            [(heading, sub) for heading, subdict in aver_data.items() for sub in subdict]
        )
        data_aver = [
            [aver_data[heading][sub][i] for heading, subdict in aver_data.items() for sub in subdict]
            for i in [0, 1]
        ]
        df_aver = pd.DataFrame(data_aver, columns=columns_aver)

        trial_file = os.path.join(self.participant_folder, f"{self.part_id}.xlsx")
        if os.path.exists(trial_file):
            os.remove(trial_file)

        df_aver.index = ['LEFT', 'RIGHT']

        wb = openpyxl.Workbook()
        bold_font = Font(bold=True)

        # summary-sheet
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        trial_cell = ws_summary.cell(row=2, column=1, value='Trial')
        trial_cell.font = bold_font
        ws_summary.cell(row=2, column=1, value='Trial')
        col = 2
        for level0, level1 in df_sum.columns:
            level0_cell = ws_summary.cell(row=1, column=col, value=level0)
            level0_cell.font = bold_font
            level1_cell = ws_summary.cell(row=2, column=col, value=level1)
            level1_cell.font = bold_font
            col += 1
        for row_idx, (trial_num, row_data) in enumerate(df_sum.iterrows()):
            excel_row = row_idx + 3
            trial_num = valid_ranges[row_idx] + 1
            ws_summary.cell(row=excel_row, column=1, value=trial_num)
            for col_idx, value in enumerate(row_data, start=2):
                ws_summary.cell(row=excel_row, column=col_idx, value=value)

        # average-sheet
        ws_average = wb.create_sheet('Average')
        bh_cell = ws_average.cell(row=2, column=1, value='BH')
        bh_cell.font = bold_font
        col = 2
        for level0, level1 in df_aver.columns:
            level0_cell = ws_average.cell(row=1, column=col, value=level0)
            level0_cell.font = bold_font
            level1_cell = ws_average.cell(row=2, column=col, value=level1)
            level1_cell.font = bold_font
            col += 1
        bh_values = ['LEFT', 'RIGHT']
        for row_idx, (bh_value, row_data) in enumerate(df_aver.iterrows()):
            excel_row = row_idx + 3
            bh_label = bh_values[row_idx]
            ws_average.cell(row=excel_row, column=1, value=bh_label)
            for col_idx, value in enumerate(row_data, start=2):
                ws_average.cell(row=excel_row, column=col_idx, value=value)

        current_col = 2
        current_header = None
        start_col = 2

        for level0, level1 in df_sum.columns:
            if level0 != current_header:
                if current_header is not None and current_col > start_col:
                    ws_summary.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=current_col - 1)
                    ws_summary.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

                current_header = level0
                start_col = current_col
            current_col += 1

        if current_col > start_col:
            ws_summary.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=current_col - 1)
            ws_summary.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

        current_col = 2
        current_header = None
        start_col = 2

        for level0, level1 in df_aver.columns:
            if level0 != current_header:
                if current_header is not None and current_col > start_col:
                    ws_average.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=current_col - 1)
                    ws_average.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

                current_header = level0
                start_col = current_col
            current_col += 1

        if current_col > start_col:
            ws_average.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=current_col - 1)
            ws_average.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

        wb.save(trial_file)

