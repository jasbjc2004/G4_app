import os
from copy import copy
from pathlib import Path
import re

import openpyxl
import pandas as pd
from PySide6.QtCore import QObject, Signal, QThread
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (
    QVBoxLayout, QWidget, QPushButton, QFileDialog, QApplication, QDialog, QMessageBox
)
from openpyxl.styles import Font, Alignment

from constants import BIMAN_PARAMS, UNIMAN_PARAMS
from logger import get_logbook
from widget_settings import manage_settings
from data_processing import calculate_extra_parameters, predict_score, calculate_boxhand

from window_set_up import SetUp


class StartUp(QDialog):
    """
    Window at the beginning of the program, to load existing project or start a new one
    """
    def __init__(self):
        super().__init__()
        self.thread = None
        self.worker = None
        self.progression = None
        self.logger = get_logbook('window_start_up')
        self.setWindowTitle("Startup")
        file_directory = (os.path.dirname(os.path.abspath(__file__)))
        dir_icon = os.path.join(file_directory, 'NEEDED/PICTURES/hands.ico')
        self.setWindowIcon(QIcon(dir_icon))

        layout = QVBoxLayout()

        self.button_new = QPushButton("New Project")
        self.button_new.clicked.connect(self.open_setup)
        self.button_old = QPushButton("Open existing Project")
        self.button_old.clicked.connect(self.reopen_setup)
        self.compare_patients = QPushButton("Compare participants")
        self.compare_patients.clicked.connect(self.select_patients)

        layout.addWidget(self.button_new)
        layout.addWidget(self.button_old)
        layout.addWidget(self.compare_patients)

        self.setLayout(layout)

    # go to setup window
    def open_setup(self):
        self.setup = SetUp()
        self.setup.show()
        self.close()

    def reopen_setup(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Trial Directory", os.path.expanduser("~/Documents"))
        if folder:
            self.setup = SetUp(folder)
            self.setup.show()
            self.close()

    def select_patients(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Patient Directory", os.path.expanduser("~/Documents"))
        if folder:
            try:
                self.make_progress()
                self.thread = QThread()
                self.worker = CompareWorker(folder)
                self.worker.moveToThread(self.thread)

                self.worker.progression.connect(self.set_progress)
                self.worker.done.connect(self.finish)

                self.thread.started.connect(self.worker.run)
                self.thread.start()
            except Exception as e:
                QMessageBox.critical(self, 'Error', str(e))
                
    def make_progress(self):
        """
        Make the progress bar pop-up
        """
        from widget_progression_bar import ProgressionBar

        self.progression = ProgressionBar()
        self.progression.show()

    def finish(self):
        """
        Finishing of the comparison between patients.
        :return:
        """
        if self.progression:
            self.progression.set_progress(100)

        QMessageBox.information(self, "Success", "Successfully compared between patients!")

        self.worker.done.connect(self.thread.quit)
        self.worker.done.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)

    def set_progress(self, value: int):
        self.progression.set_progress(value)

    def show_error(self, e):
        self.logger.error(e, exc_info=True)
        if self.progression:
            self.progression.close()
            self.progression = None
        QMessageBox.critical(self, "Export Error", f"An error occurred during comparison: {str(e)}")


class CompareWorker(QThread):
    """
    Thread to compare the patients --> possible to update the progress bar without freezing the program
    """
    progression = Signal(int)
    done = Signal()
    error = Signal(str)

    def __init__(self, folder):
        super().__init__()
        self.folder = folder
        self.interval = 100 / (sum(1 for p in Path(folder).iterdir() if p.is_dir())+1)
        self.counter = 0

    def run(self):
        self.progression.emit(0)
        try:
            self.search_dir(self.folder)
        except Exception as e:
            self.error.emit(str(e))

    def add_patient_data(self, file, data_dict, aver_data, trials):
        """
        Add all necessary data of the patients trial to the dict
        :param file: the file containing the data of the patient
        :param data_dict: dict containing data of the patient
        :param aver_data: dict containing average of all the patient
        :param trials: list of the trials of a patient
        :return:
        """
        trial_data = pd.read_excel(file)
        trial_number = file.name.split('.')[-2].split('_')[-1]
        NUMBER_EVENTS = manage_settings.get("Events", "NUMBER_EVENTS")

        first_col = trial_data.iloc[:, 0]

        if trial_data.shape[1] < 1 or first_col.isnull().all() or first_col.astype(str).str.strip().eq("").all():
            return

        xs = trial_data.iloc[:, 0].values
        if len(xs) < 2:
            return

        print('starting appending data')

        trials.append(int(trial_number)-1)

        log_left, log_right = [], []

        x1 = trial_data.iloc[:, 1].values
        y1 = trial_data.iloc[:, 2].values
        z1 = trial_data.iloc[:, 3].values
        v1 = trial_data.iloc[:, 4].values

        x2 = trial_data.iloc[:, 5].values
        y2 = trial_data.iloc[:, 6].values
        z2 = trial_data.iloc[:, 7].values
        v2 = trial_data.iloc[:, 8].values

        for i in range(len(x1)):
            log_left.append((x1[i], y1[i], z1[i], v1[i],))
            log_right.append((x2[i], y2[i], z2[i], v2[i],))

        if trial_data.shape[1] < 11:
            event_log = [0] * NUMBER_EVENTS
        elif True:  # eventueel toevoegen in instellingen of extra pop-up (manual events)
            event_log = trial_data.iloc[:, 12].values[0:NUMBER_EVENTS].tolist()
            if all(x == 0 for x in event_log):
                event_log = trial_data.iloc[:, 11].values[0:NUMBER_EVENTS].tolist()
        else:
            event_log = trial_data.iloc[:, 11].values[0:NUMBER_EVENTS].tolist()
        event_log = [int(ei) for ei in event_log[:]]

        print(log_left, log_right)
        score = predict_score(log_left, log_right)
        case = calculate_boxhand(log_left, log_right, score)
        print(event_log, score, case)
        if case == 0:
            bim_par, uni_par = calculate_extra_parameters(event_log, log_right, log_left)
        elif case == 1:
            bim_par, uni_par = calculate_extra_parameters(event_log, log_left, log_right)
        else:
            bim_par, uni_par = [0] * len(data_dict["Bimanual"].keys()), [0] * len(data_dict["Unimanual"].keys())

        print(event_log, bim_par, uni_par)

        data_dict[" "]["Score"].append(score)
        data_dict[" "][""].append('')
        data_dict[" "][" "].append('')

        for key, value in zip(data_dict["Events"].keys(), event_log):
            data_dict["Events"][key].append(value)

        for key, value in zip(data_dict["Bimanual"].keys(), bim_par):
            data_dict["Bimanual"][key].append(value)
            if case == 0:
                aver_data["Bimanual"][key][0] += value
            elif case == 1:
                aver_data["Bimanual"][key][1] += value

        if case == 0:
            aver_data[""]["Number of Trials"][0] += 1
        elif case == 1:
            aver_data[""]["Number of Trials"][1] += 1

        for key, value in zip(data_dict["Unimanual"].keys(), uni_par):
            data_dict["Unimanual"][key].append(value)

            if case == 0:
                aver_data["Unimanual"][key][0] += value
            elif case == 1:
                aver_data["Unimanual"][key][1] += value

        print(data_dict)

    def add_data_aver(self, aver_data, wb):
        print(aver_data)
        for param in aver_data["Bimanual"]:
            aver_data["Bimanual"][param] = [param_lr / aver_data[""]["Number of Trials"][index]
                                            if aver_data[""]["Number of Trials"][index] != 0 else 0
                                            for index, param_lr in enumerate(aver_data["Bimanual"][param])]

        for param in aver_data["Unimanual"]:
            aver_data["Unimanual"][param] = [param_lr / aver_data[""]["Number of Trials"][index]
                                             if aver_data[""]["Number of Trials"][index] != 0 else 0
                                             for index, param_lr in enumerate(aver_data["Unimanual"][param])]

        columns_aver = pd.MultiIndex.from_tuples(
            [(heading, sub) for heading, subdict in aver_data.items() for sub in subdict]
        )
        data_aver = [
            [aver_data[heading][sub][i] for heading, subdict in aver_data.items() for sub in subdict]
            for i in [0, 1]
        ]
        df_aver = pd.DataFrame(data_aver, columns=columns_aver)
        df_aver.index = ['LEFT', 'RIGHT']

        bold_font = Font(bold=True)

        # average info
        print('hope')
        ws_average = wb[wb.sheetnames[0]]
        ws_average.title = "Comparison"
        for merged_range in list(ws_average.merged_cells.ranges):
            ws_average.unmerge_cells(str(merged_range))
        bh_cell = ws_average.cell(row=2, column=1, value='BH')
        bh_cell.font = bold_font
        col = 2
        print('fully')
        for level0, level1 in df_aver.columns:
            level0_cell = ws_average.cell(row=1, column=col, value=level0)
            level0_cell.font = bold_font
            level1_cell = ws_average.cell(row=2, column=col, value=level1)
            level1_cell.font = bold_font
            col += 1
        bh_values = ['LEFT', 'RIGHT']
        print('almost')
        for row_idx, (bh_value, row_data) in enumerate(df_aver.iterrows()):
            excel_row = row_idx + 3
            bh_label = bh_values[row_idx]
            ws_average.cell(row=excel_row, column=1, value=bh_label)
            for col_idx, value in enumerate(row_data, start=2):
                ws_average.cell(row=excel_row, column=col_idx, value=value)

        print('maybe')

        current_col = 2
        current_header = None
        start_col = 2

        for level0, level1 in df_aver.columns:
            if level0 != current_header:
                if current_header is not None and current_col > start_col:
                    ws_average.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=current_col - 1)
                    ws_average.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

                current_header = level0
                start_col = current_col
            current_col += 1

        if current_col > start_col:
            ws_average.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=current_col - 1)
            ws_average.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

    def add_data_sum(self, part_code, valid_ranges, sum_data, wb):
        """
        Add all the data to the Excel of a single participant
        :param part_code: the participant code
        :param valid_ranges: all the trials used in documents of a single participant
        :param sum_data: all the data of the participant
        :param wb: needed to add data to the Excel
        """
        columns_sum = pd.MultiIndex.from_tuples([(heading, sub) for heading, subdict in sum_data.items()
                                                 for sub in subdict])
        first_heading = next(iter(sum_data))
        first_sub = next(iter(sum_data[first_heading]))
        n_rows = len(sum_data[first_heading][first_sub])
        data_sum = [
            [sum_data[heading][sub][i] for heading, subdict in sum_data.items() for sub in subdict]
            for i in range(n_rows)
        ]
        df_sum = pd.DataFrame(data_sum, columns=columns_sum)

        bold_font = Font(bold=True)

        ws_summary = wb.create_sheet(part_code)
        trial_cell = ws_summary.cell(row=2, column=1, value='Trial')
        trial_cell.font = bold_font
        ws_summary.cell(row=2, column=1, value='Trial')
        col = 2
        for level0, level1 in df_sum.columns:
            level0_cell = ws_summary.cell(row=1, column=col, value=level0)
            level0_cell.font = bold_font
            level1_cell = ws_summary.cell(row=2, column=col, value=level1)
            level1_cell.font = bold_font
            col += 1
        for row_idx, (trial_num, row_data) in enumerate(df_sum.iterrows()):
            excel_row = row_idx + 3
            trial_num = valid_ranges[row_idx] + 1
            ws_summary.cell(row=excel_row, column=1, value=trial_num)
            for col_idx, value in enumerate(row_data, start=2):
                ws_summary.cell(row=excel_row, column=col_idx, value=value)

        current_col = 2
        current_header = None
        start_col = 2

        for level0, level1 in df_sum.columns:
            if level0 != current_header:
                if current_header is not None and current_col > start_col:
                    ws_summary.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=current_col - 1)
                    ws_summary.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

                current_header = level0
                start_col = current_col
            current_col += 1

        if current_col > start_col:
            ws_summary.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=current_col - 1)
            ws_summary.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

    def search_dir(self, folder):
        """
        Search inside the folder for all necessary files
        """
        compare_file = os.path.join(folder, f"Compare_patients.xlsx")
        if os.path.exists(compare_file):
            os.remove(compare_file)

        wb_dest = openpyxl.Workbook()

        aver_data = {"": {"Number of Trials": [0, 0]}, "Bimanual": {}, "Unimanual": {}}
        for param in BIMAN_PARAMS:
            aver_data["Bimanual"][param] = [0, 0]
        for param in UNIMAN_PARAMS:
            aver_data["Unimanual"][param] = [0, 0]

        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)

            if os.path.isdir(file_path):
                part_code = filename
                if '(' in filename and ')' in filename:
                    part_code = filename.split('(')[0]

                """
                if part_code+'.xlsx' in os.listdir(file_path):
                    print(file_path)

                    wb_src = openpyxl.load_workbook(os.path.join(file_path, part_code+'.xlsx'))
                    wb_src = wb_src.active

                    wb_part = wb_dest.create_sheet(part_code)

                    col_par = -1
                    for row in wb_src.iter_rows():
                        for cell in row:
                            new_cell = wb_part.cell(row=cell.row, column=cell.column, value=cell.value)
                            if cell.value == BIMAN_PARAMS[0]:
                                col_par = cell.column
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)

                        for merged_range in wb_src.merged_cells.ranges:
                            wb_part.merge_cells(str(merged_range))

                        for col_letter, dim in wb_src.column_dimensions.items():
                            wb_part.column_dimension[col_letter].width = dim.width

                        for row_num, dim in wb_src.row_dimensions.items():
                            wb_part.row_dimension[row_num].height = dim.height
                """

                LABEL_EVENT = manage_settings.get("Events", "LABEL_EVENT")

                sum_data = {
                    "Events": {}
                }
                for key in LABEL_EVENT:
                    sum_data["Events"][key] = []

                sum_data[" "] = {}
                sum_data[" "][" "] = []
                sum_data[" "]["Score"] = []
                sum_data[" "][""] = []

                sum_data["Bimanual"] = {}
                for param in BIMAN_PARAMS:
                    sum_data["Bimanual"][param] = []

                sum_data["Unimanual"] = {}
                for param in UNIMAN_PARAMS:
                    sum_data["Unimanual"][param] = []

                def get_index(file_trial):
                    match = re.search(r'trial_(\d+)', file_trial.stem)
                    return int(match.group(1)) if match else float('inf')

                print('here fine')

                files = sorted(Path(file_path).glob("trial_*.xlsx"), key=get_index)

                print(files)

                trial_number = []
                for file in files:
                    # print(f'starting file: {file}')
                    self.add_patient_data(file, sum_data, aver_data, trial_number)
                    # print(f'done file: {file}')

                print('done')
                print(sum_data)

                self.add_data_sum(part_code, trial_number, sum_data, wb_dest)

                self.counter += 1
                self.progression.emit(self.counter*self.interval)

        self.add_data_aver(aver_data, wb_dest)
        wb_dest.save(compare_file)
        self.done.emit()
